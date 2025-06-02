import tkinter as hansdev
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from PIL import Image, ImageTk
import os

accdatabase = "accounts.xlsx"
recorddatabase ="record.xlsx"
background = "bg.jpg"

def open_main():
    def checkrecorddb():
        if not os.path.exists(recorddatabase):
            wb = Workbook()
            ws = wb.active
            ws.append(["No.", "Lastname", "Firstname", "Initial", "Course", "Section", "Violence", "Date"])
            wb.save(recorddatabase)
        
    def load_data():
        database = load_workbook(recorddatabase)
        sheet = database.active
    
        list_values = list(sheet.values)
        print(list_values)
        for col_name in list_values[0]:
            exelview.heading(col_name, text=col_name)
        for student_value in list_values[1:]:
            exelview.insert('', hansdev.END, values=student_value)
    
    def add_record():
        
        def insert_student():
            number = student_number_entry.get()
            lastname = student_lastname_entry.get()
            firstname = student_firstname_entry.get()
            middle = student_middle_entry.get()
            course = student_course_entry.get()
            section = student_section_entry.get()
            violence = student_type_entry.get()
            date = student_date_entry.get()
            
            
            database = load_workbook(recorddatabase)
            sheet = database.active
            row_values = [number, lastname, firstname, middle, course, section, violence, date]
            sheet.append(row_values)
            database.save(recorddatabase)
            
            exelview.insert('', hansdev.END, values=row_values)
        
        def others():
            if student_type_entry.get() == "Others":
               student_others_label = hansdev.Label(student_viol_frame, text="Others")
               student_others_label.grid(row=0, column=2)
               student_others_entry = hansdev.Entry(student_info_frame)
               student_others_entry.grid(row=1, column=2)
        
        
        def enter_data():
            if student_number_entry.get() == "": 
                messagebox.showerror(title="Data", message="Input the student number.")
                #others()
            elif student_firstname_entry.get() == "": 
                messagebox.showerror(title="Data", message="Input the student real name.")
            elif student_lastname_entry.get() == "":
                messagebox.showerror(title="Data", message="Input the student lastname.")
            elif student_middle_entry.get() == "--Select--":
                messagebox.showerror(title="Data", message="Input the student middle initial.")
            elif student_course_entry.get() == "--Select--":
                messagebox.showerror(title="Data", message="Input the student course.")
            elif student_section_entry.get() == "--Select--":
                messagebox.showerror(title="Data", message="Input the student section.")
            elif student_type_entry.get() == "--Select--":
                messagebox.showerror(title="Data", message="Input the student violence.")
            elif student_date_entry.get() == "Month/Day/Year":
                messagebox.showerror(title="Data", message="Input the student date of violence.")
            else:
                messagebox.showinfo("Record Has Been Successfully", "Updated!")
                insert_student()
                record.destroy()
 
        record = hansdev.Tk()
        record.title("Student Violence Log")
        record_frame = ttk.Frame(record)
        record_frame.pack()
        
        #student detailes input
        student_info_frame = hansdev.LabelFrame(record_frame, text="Student Information")
        student_info_frame.grid(row=0, column=0, sticky="news", padx=20, pady=20)
        
        student_number_label = hansdev.Label(student_info_frame, text="Student No.")
        student_number_label.grid(row=0, column=0)
        student_lastname_label = hansdev.Label(student_info_frame, text="Last Name")
        student_lastname_label.grid(row=0, column=1)
        student_firstname_label = hansdev.Label(student_info_frame, text="First Name")
        student_firstname_label.grid(row=0, column=2)
        student_middle_label = hansdev.Label(student_info_frame, text="Middle Name")
        student_middle_label.grid(row=2, column=0)
        student_course_label = hansdev.Label(student_info_frame, text="Course")
        student_course_label.grid(row=2, column=1)
        student_section_label = hansdev.Label(student_info_frame, text="Section")
        student_section_label.grid(row=2, column=2)
        
        
        student_number_entry = hansdev.Entry(student_info_frame)
        student_lastname_entry = hansdev.Entry(student_info_frame)
        student_firstname_entry = hansdev.Entry(student_info_frame)
        student_middle_entry = ttk.Combobox(student_info_frame, values=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"])
        student_course_entry = ttk.Combobox(student_info_frame, values=["BSCS", "BSMA", "BPA", "BSCRIM"])
        student_section_entry = ttk.Combobox(student_info_frame, values=["1A", "1B", "1C", "1D", "1E", "1F"])
        student_middle_entry.set("--Select--")
        student_course_entry.set("--Select--")
        student_section_entry.set("--Select--")
        student_number_entry.grid(row=1, column=0)
        student_lastname_entry.grid(row=1, column=1)
        student_firstname_entry.grid(row=1, column=2)
        student_middle_entry.grid(row=3, column=0)
        student_course_entry.grid(row=3, column=1)
        student_section_entry.grid(row=3, column=2)
        
        for widget in student_info_frame.winfo_children():
            widget.grid_configure(padx=7, pady=5)
        
        #student violation input
        student_viol_frame = hansdev.LabelFrame(record_frame, text="Student Violation")
        student_viol_frame.grid(row=1, column=0, sticky="news", padx=20, pady=20)
        
        student_type_label = hansdev.Label(student_viol_frame, text="Violation Type")
        student_type_label.grid(row=0, column=0)
        student_date_label = hansdev.Label(student_viol_frame, text="Date and Time")
        student_date_label.grid(row=0, column=1)
        
        student_type_entry = ttk.Combobox(student_viol_frame, values=["Late", "No ID", "Piercing", "Missing Tie", "Wrong Attire", "Improper Shoes", "Safety Violations", "Property Violations", "Skipping Violations", "Sexual Misconduct", "Technology Misuse", "Academic Violations", "Behavioral Violations", "Substance-Related Violations", "Others"])
        student_type_entry.set("--Select--")
        dates = [(datetime.today() + timedelta(days=i)).strftime("%m/%d/%Y") for i in range(30)]
        student_date_entry = ttk.Combobox(student_viol_frame, values=dates)
        student_date_entry.set("Month/Day/Year")
        student_type_entry.grid(row=1, column=0)
        student_date_entry.grid(row=1, column=1)
        
        
        for widget in student_viol_frame.winfo_children():
            widget.grid_configure(padx=7, pady=5)
            
        enter = ttk.Button(record_frame, text="Enter Data", command=enter_data)
        enter.grid(row=2, column=0, sticky="news", pady=20, padx=20)
        
        record.mainloop()
    
    def delete_record():
        selected = exelview.selection()
        if not selected:
            messagebox.showerror(title="No Selected", message="Select that you want to delete.")
            return
        
        selected_values = exelview.item(selected[0], 'values')
        selected_values = tuple(str(v) for v in selected_values)
        database = load_workbook(recorddatabase)
        sheet = database.active
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_values = tuple(str(cell.value) for cell in row)
            if row_values == selected_values:
                sheet.delete_rows(idx)
                break
        database.save(recorddatabase)
        exelview.delete(*exelview.get_children())
        load_data()
        messagebox.showinfo("Record Has Been Successfully", "Updated!")
        
    def update_record():
        selected = exelview.selection()
        if not selected:
            messagebox.showerror(title="No Selection", message="Please select a record to update.")
            return

        old_values = exelview.item(selected[0], 'values')

        update_window = hansdev.Toplevel()
        update_window.title("Update Record")

        form_frame = ttk.LabelFrame(update_window, text="Edit Student Record")
        form_frame.pack(padx=10, pady=10)

        labels = ["Student No.", "Last Name", "First Name", "Middle", "Course", "Section", "Violation", "Date"]
        fields = []

        for i, label in enumerate(labels):
            ttk.Label(form_frame, text=label).grid(row=i, column=0, sticky="w", pady=3)

            if label == "Middle":
                entry = ttk.Combobox(form_frame, values=["--Select--"] + list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
            elif label == "Course":
                entry = ttk.Combobox(form_frame, values=["--Select--", "BSCS", "BSMA", "BPA", "BSCRIM"])
            elif label == "Section":
                entry = ttk.Combobox(form_frame, values=["--Select--", "1A", "1B", "1C", "1D", "1E", "1F"])
            elif label == "Violation":
                entry = ttk.Combobox(form_frame, values=["--Select--", "Late", "No ID", "Piercing", "Missing Tie", "Wrong Attire", "Improper Shoes", "Safety Violations", "Property Violations", "Skipping Violations", "Sexual Misconduct", "Technology Misuse", "Academic Violations", "Behavioral Violations", "Substance-Related Violations", "Others"])
            elif label == "Date":
                dates = [(datetime.today() + timedelta(days=i)).strftime("%m/%d/%Y") for i in range(30)]
                entry = ttk.Combobox(form_frame, values=dates)
            else:
                entry = hansdev.Entry(form_frame)

            entry.grid(row=i, column=1, pady=3, padx=5)
            entry.insert(0, old_values[i])
            fields.append(entry)

        def save_update():
            new_values = [entry.get() for entry in fields]

            # Update Excel
            database = load_workbook(recorddatabase)
            sheet = database.active
            for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_data = [str(cell.value) for cell in row]
                if tuple(row_data) == old_values:
                    for col, value in enumerate(new_values, start=1):
                        sheet.cell(row=idx, column=col).value = value
                    break
            database.save(recorddatabase)

            exelview.delete(*exelview.get_children())
            load_data()
            messagebox.showinfo("Update Successful", "The record has been updated.")
            update_window.destroy()

        save_btn = ttk.Button(update_window, text="Save", command=save_update)
        save_btn.pack(pady=10)
        
    checkrecorddb()
        
    home = hansdev.Tk()
    home.title("Security Panel")
    home_box = ttk.Frame(home)
    home_box.pack()
    
    security_panel = ttk.Labelframe(home_box, text="Admnistrator")
    security_panel.grid(row=0, column=0, padx=20, pady=10)
    
    insert = ttk.Button(security_panel, text="Insert", command=add_record)
    insert.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="nsew")
    insert = ttk.Button(security_panel, text="Update", command=update_record)
    insert.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="nsew")
    insert = ttk.Button(security_panel, text="Delete", command=delete_record)
    insert.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="nsew")
    
    #admin_profile = ttk.Label(security_panel, text="  🖥️ ", font=("Arial", 80))
    #admin_profile.grid(row=0, column=0)

    #home
    #home_frame_0 = ttk.LabelFrame(home_box)
    #home_frame_0.grid(row=0, column=0, sticky="news")
    #add = ttk.Button(home, text="Login", command=add_record)
    #add.grid(row=0, column=0, sticky="news", pady=10, padx=30)
    
    exel_frame = ttk.Frame(home_box)
    exel_frame.grid(row=0, column=1, pady=10)
    exel_scroll = ttk.Scrollbar(exel_frame)
    exel_scroll.pack(side="right", fill="y")
    
    head = ("No.", "Lastname", "Firstname", "Initial", "Course", "Section", "Violence", "Date")
    exelview = ttk.Treeview(exel_frame, show="headings",
                            yscrollcommand=exel_scroll.set, columns=head, height=13)
    exelview.column("No.", width=75)
    exelview.column("Lastname", width=100)
    exelview.column("Firstname", width=100)
    exelview.column("Initial", width=60)
    exelview.column("Course", width=60)
    exelview.column("Section", width=60)
    exelview.column("Violence", width=100)
    exelview.column("Date", width=100)
    exelview.pack()
    exel_scroll.config(command=exelview.yview)
    load_data()
    
    home.mainloop()
    
# nagawa ng file para sa acc
def checkaccdb():
    if not os.path.exists(accdatabase):
        wb = Workbook()
        ws = wb.active
        ws.append(["Username", "Password"])
        wb.save(accdatabase)

def add_user(username, password):
    wb = load_workbook(accdatabase)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return False
    ws.append([username, password])
    wb.save(accdatabase)
    return True

def validate_login(username, password):
    wb = load_workbook(accdatabase)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

def find_password_by_username(username):
    wb = load_workbook(accdatabase)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return row[1]
    return None

# Pag pinindot ang login
def login_button():  
    accepted = accept_var.get()
    if accepted=="Accepted":
        username = label_user_input.get()
        password = label_pass_input.get()
        if not username or not password:
            messagebox.showerror("Error", "Please enter username and password.")
            return
        if validate_login(username, password):
            messagebox.showinfo("Login Successful", f"Welcome, {username}!")
            window.destroy()
            open_main()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password.")
    else:
        messagebox.showerror(title="Not Autorized", message="You are not autorized please try again later.")
    

# Pag pinindot ang Signup
def signup_button():
    username = label_user_input.get()
    password = label_pass_input.get()
    if not username or not password:
        messagebox.showerror("Error", "Please fill both username and password to signup.")
        return
    if add_user(username, password):
        messagebox.showinfo("Signup Success", "User registered successfully! You can login now.")
    else:
        messagebox.showerror("Signup Failed", "Username already exists.")

def forget_password():
    window = hansdev.Toplevel()
    window.title("Forget Password")
    window.geometry("300x150")

    ttk.Label(window, text="Enter your username:").pack(pady=10)

    username_entry = ttk.Entry(window)
    username_entry.pack(pady=5)

    def findpass():
        username = username_entry.get()
        if username:
            pwd = find_password_by_username(username)
            if pwd:
                messagebox.showinfo("Password Found", f"Your password is: {pwd}", parent=window)
            else:
                messagebox.showerror("Not Found", "Username not registered.", parent=window)
        else:
            messagebox.showwarning("Input Error", "Please enter a username.", parent=window)

    submit_btn = ttk.Button(window, text="Find ur password.", command=findpass)
    submit_btn.pack(pady=10)

    username_entry.focus()
    window.grab_set()

checkaccdb()

window = hansdev.Tk()
window.title("Login System")
window.geometry("900x600")
window.resizable(False, False)

# indicates bg iamage
bg_image = Image.open(background)
bg_image = bg_image.resize((900, 600), Image.Resampling.LANCZOS)
bg_photo = ImageTk.PhotoImage(bg_image)

background_label = hansdev.Label(window, image=bg_photo)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

# Transparent centered
main_frame = hansdev.Frame(window, bg=None, highlightthickness=0)
main_frame.place(relx=0.5, rely=0.5, anchor="center")

main_box = ttk.Frame(main_frame)
main_box.grid(sticky="news")

main_label_0 = ttk.Label(main_box)
main_label_0.grid(row=0, column=0, sticky="news")

login_label = ttk.Label(main_label_0, text="Security," \
" Dashboard", font=("Arial", 25))
login_label.grid(row=0, column=0, sticky="news", pady=30, padx=30)
label_user = ttk.Label(main_label_0, text="Username", font=("Arial", 12))
label_user_input = ttk.Entry(main_label_0, font=("Arial", 12))
label_user_input.insert(0, "example garciahxns@gmail.com")
label_user_input.bind("<FocusIn>", lambda tanggal: label_user_input.delete('0', 'end'))
label_pass = ttk.Label(main_label_0, text="Password", font=("Arial", 12))
label_pass_input = ttk.Entry(main_label_0, show="*", font=("Arial", 12))
#label_noacc = ttk.Label(main_label_0, text="You dont know the admin account?", font=("Arial", 8))
label_user.grid(row=1, column=0, sticky="news", padx=30, pady=4)
forget = hansdev.Button(main_label_0, text="Forgot Account?", font=("Arial", 8), command=forget_password)
label_user_input.grid(row=2, column=0, sticky="news", padx=30)
label_pass.grid(row=3, column=0, sticky="news", padx=30, pady=4)
label_pass_input.grid(row=4, column=0, sticky="news", padx=30)
#label_noacc.grid(row=5, column=0, sticky="e" , padx=40, pady=4)
forget.grid(row=5, column=0, sticky="e" , padx=40, pady=4)


autorize_frame = ttk.LabelFrame(main_label_0, text="Are you autorize personel?")
autorize_frame.grid(row=6, column=0, sticky="news", pady=15, padx=30)
accept_var = hansdev.StringVar(value="Not Accepted")
autorize_check = ttk.Checkbutton(autorize_frame, text="Check if you are autorized personel.",
                                      variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
autorize_check.grid(row=0, column=0, sticky="news")
main_label_1 = ttk.Label(main_box)
main_label_1.grid(row=0, column=1, sticky="news")

signlog = ttk.Frame(main_label_0)
signlog.grid(row=7, column=0, sticky="ew", pady=10, padx=30)

signlog.columnconfigure(0, weight=1) # Allow columns to expand equally
signlog.columnconfigure(1, weight=1)

login = ttk.Button(signlog, text="Login", command=login_button)
signup = ttk.Button(signlog, text="Signup", command=signup_button)

login.grid(row=0, column=1, padx=2, sticky="nsew")
signup.grid(row=0, column=0, padx=2, sticky="nsew")

last = ttk.LabelFrame(main_label_0, width=10)
last.grid(row=8, column=0, sticky="news", padx=10, pady=10)

window.mainloop()
